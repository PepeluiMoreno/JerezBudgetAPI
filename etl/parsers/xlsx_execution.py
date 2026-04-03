"""
Parser de XLSX de ejecución presupuestaria del Ayuntamiento de Jerez.

Estrategia de detección de cabecera:
  - Busca en las primeras 25 filas la que contenga palabras clave
    ("orgánica", "económica", "crédito", "previsión") en >= 3 columnas.
  - No asume posiciones fijas — mapea columnas por nombre normalizado.
  - Esto hace el parser robusto ante cambios de formato entre ejercicios.

Salida: lista de dicts con campos normalizados listos para el ORM.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

import openpyxl
import structlog

logger = structlog.get_logger(__name__)

# ── Palabras clave para detectar cabecera ────────────────────────────────────
_HEADER_KEYWORDS = {
    "organica", "organica", "economica", "económica",
    "credito", "crédito", "prevision", "previsión",
    "denominacion", "denominación", "obligaciones", "derechos",
    "modificaciones", "pagos", "recaudacion",
}

# ── Mapeo de variantes de nombre de columna → nombre normalizado ─────────────
_COL_ALIASES: dict[str, str] = {
    # Clasificaciones
    "orgánica":                       "organic_code",
    "organica":                       "organic_code",
    "org.":                           "organic_code",
    "programa":                       "functional_code",
    "prog.":                          "functional_code",
    "económica":                      "economic_code",
    "economica":                      "economic_code",
    "ec.":                            "economic_code",
    "aplicacion":                     "economic_code",
    "aplicación":                     "economic_code",
    "denominacion":                   "description",
    "denominación":                   "description",
    "concepto":                       "description",

    # Gastos
    "créditos iniciales":             "initial_credits",
    "creditos iniciales":             "initial_credits",
    "crédito inicial":                "initial_credits",
    "credito inicial":                "initial_credits",
    "modificaciones":                 "modifications",
    "créditos definitivos":           "final_credits",
    "creditos definitivos":           "final_credits",
    "crédito definitivo":             "final_credits",
    "credito definitivo":             "final_credits",
    "autorizaciones":                 "commitments",
    "autorizaciones y disposiciones": "commitments",
    "a/d":                            "commitments",
    "obligaciones reconocidas":       "recognized_obligations",
    "obligaciones rec.":              "recognized_obligations",
    "obligaciones":                   "recognized_obligations",
    "pagos realizados":               "payments_made",
    "pagos liquid.":                  "payments_made",
    "pagos":                          "payments_made",
    "pendiente de pago":              "pending_payment",
    "pdte. pago":                     "pending_payment",

    # Ingresos
    "previsiones iniciales":          "initial_forecast",
    "previsión inicial":              "initial_forecast",
    "prevision inicial":              "initial_forecast",
    "previsiones definitivas":        "final_forecast",
    "previsión definitiva":           "final_forecast",
    "prevision definitiva":           "final_forecast",
    "derechos reconocidos netos":     "recognized_rights",
    "derechos rec. netos":            "recognized_rights",
    "derechos reconocidos":           "recognized_rights",
    "recaudación neta":               "net_collection",
    "recaudacion neta":               "net_collection",
    "recaudación":                    "net_collection",
    "recaudacion":                    "net_collection",
    "pendiente de cobro":             "pending_collection",
    "pdte. cobro":                    "pending_collection",
}

# Campos numéricos de gasto e ingreso
_EXPENSE_FIELDS = {
    "initial_credits", "modifications", "final_credits",
    "commitments", "recognized_obligations", "payments_made", "pending_payment",
}
_REVENUE_FIELDS = {
    "initial_forecast", "final_forecast",
    "recognized_rights", "net_collection", "pending_collection",
}
_NUMERIC_FIELDS = _EXPENSE_FIELDS | _REVENUE_FIELDS

# Patrón para códigos presupuestarios: secuencia de dígitos (y alguna letra)
_CODE_PATTERN = re.compile(r"^\d[\d.]*$")


@dataclass
class ParsedBudgetLine:
    """Una línea presupuestaria extraída del XLSX."""
    organic_code: str = ""
    functional_code: str = ""
    economic_code: str = ""
    description: str = ""
    direction: str = ""       # "expense" | "revenue" — inferido del contenido

    # Gastos
    initial_credits: Optional[Decimal] = None
    modifications: Optional[Decimal] = None
    final_credits: Optional[Decimal] = None
    commitments: Optional[Decimal] = None
    recognized_obligations: Optional[Decimal] = None
    payments_made: Optional[Decimal] = None
    pending_payment: Optional[Decimal] = None

    # Ingresos
    initial_forecast: Optional[Decimal] = None
    final_forecast: Optional[Decimal] = None
    recognized_rights: Optional[Decimal] = None
    net_collection: Optional[Decimal] = None
    pending_collection: Optional[Decimal] = None


@dataclass
class ParseResult:
    lines: list[ParsedBudgetLine] = field(default_factory=list)
    direction: str = "expense"        # "expense" | "revenue"
    header_row: int = 0
    column_map: dict[str, int] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    total_rows_read: int = 0
    total_rows_skipped: int = 0


# ── Utilidades ───────────────────────────────────────────────────────────────

def _normalize_header(val) -> str:
    """Normaliza un valor de celda de cabecera para el mapeo."""
    if val is None:
        return ""
    return str(val).strip().lower().replace("\n", " ").replace("  ", " ")


def _to_decimal(val) -> Optional[Decimal]:
    """Convierte un valor de celda a Decimal. None si vacío o no numérico."""
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(round(val, 2)))
    s = str(val).strip().replace("\xa0", "").replace(" ", "")
    # Formato español: 1.234.567,89 → 1234567.89
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _clean_code(val) -> str:
    if val is None:
        return ""
    raw = str(val).strip()
    # Algunos xlsx guardan el código como float (ej: 1.0 → "1")
    if re.match(r"^\d+\.0$", raw):
        raw = raw[:-2]
    return raw


def _is_subtotal_row(row_values: list) -> bool:
    """
    Detecta filas de subtotales/totales que no son líneas presupuestarias reales.
    Heurística: la descripción contiene palabras como "total", "suma", "capítulo".
    """
    for val in row_values:
        if val is None:
            continue
        s = str(val).lower()
        if any(kw in s for kw in ("total", "suma", "capítulo", "capitulo", "artículo", "articulo")):
            return True
    return False


# ── Detección de cabecera ────────────────────────────────────────────────────

def _find_header_row(ws) -> Optional[int]:
    """
    Busca la fila de cabecera en las primeras 25 filas.
    Retorna el índice 1-based de la fila, o None si no se encuentra.
    """
    for row_idx in range(1, 26):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, 20)]
        normalized = [_normalize_header(v) for v in row]
        matches = sum(
            1 for n in normalized
            if any(kw in n for kw in _HEADER_KEYWORDS)
        )
        if matches >= 3:
            logger.debug("header_found", row=row_idx, matches=matches)
            return row_idx
    return None


def _build_column_map(ws, header_row: int) -> dict[str, int]:
    """
    Construye el mapa {nombre_normalizado → índice_columna_1based}
    a partir de la fila de cabecera.
    """
    col_map: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        raw = ws.cell(row=header_row, column=col_idx).value
        norm = _normalize_header(raw)
        if not norm:
            continue
        # Buscar alias exacto primero
        canonical = _COL_ALIASES.get(norm)
        if not canonical:
            # Buscar alias parcial
            for alias, can in _COL_ALIASES.items():
                if alias in norm or norm in alias:
                    canonical = can
                    break
        if canonical and canonical not in col_map:
            col_map[canonical] = col_idx
            logger.debug("column_mapped", norm=norm, canonical=canonical, col=col_idx)

    return col_map


def _detect_direction(col_map: dict[str, int]) -> str:
    """Infiere si el XLSX es de gastos o ingresos por los campos presentes."""
    expense_hits = sum(1 for f in _EXPENSE_FIELDS if f in col_map)
    revenue_hits = sum(1 for f in _REVENUE_FIELDS if f in col_map)
    return "revenue" if revenue_hits > expense_hits else "expense"


# ── Parser principal ─────────────────────────────────────────────────────────

def parse_execution_xlsx(path: Path, hint_direction: Optional[str] = None) -> ParseResult:
    """
    Parsea un XLSX de ejecución presupuestaria.

    Args:
        path: Ruta al fichero .xlsx
        hint_direction: "expense" | "revenue" | None (auto-detect)

    Returns:
        ParseResult con las líneas extraídas y metadatos del parse.
    """
    result = ParseResult()

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        logger.error("xlsx_open_error", path=str(path), error=str(e))
        result.warnings.append(f"No se pudo abrir el fichero: {e}")
        return result

    # Usar primera hoja (o la que contenga datos presupuestarios)
    ws = wb.active
    if ws is None:
        ws = wb.worksheets[0]

    # Detectar cabecera
    header_row = _find_header_row(ws)
    if header_row is None:
        msg = f"No se encontró fila de cabecera en {path.name}"
        logger.warning("no_header_found", path=str(path))
        result.warnings.append(msg)
        wb.close()
        return result

    result.header_row = header_row
    col_map = _build_column_map(ws, header_row)
    result.column_map = col_map

    if not col_map:
        result.warnings.append("No se pudieron mapear columnas de la cabecera")
        wb.close()
        return result

    direction = hint_direction or _detect_direction(col_map)
    result.direction = direction

    logger.info(
        "xlsx_parse_start",
        file=path.name,
        direction=direction,
        header_row=header_row,
        mapped_cols=list(col_map.keys()),
    )

    # ── Leer filas de datos ──────────────────────────────────────────────────
    def cell(row_idx: int, field_name: str):
        col = col_map.get(field_name)
        if col is None:
            return None
        return ws.cell(row=row_idx, column=col).value

    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Leer todos los valores de la fila para heurísticas
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, min(15, ws.max_column + 1))]

        # Saltar filas completamente vacías
        if all(v is None or str(v).strip() == "" for v in row_vals):
            result.total_rows_skipped += 1
            continue

        # Saltar subtotales
        if _is_subtotal_row(row_vals):
            result.total_rows_skipped += 1
            continue

        economic_code = _clean_code(cell(row_idx, "economic_code"))

        # Saltar filas sin código económico válido
        if not economic_code or not re.match(r"^\d", economic_code):
            result.total_rows_skipped += 1
            continue

        line = ParsedBudgetLine(
            organic_code=_clean_code(cell(row_idx, "organic_code")),
            functional_code=_clean_code(cell(row_idx, "functional_code")),
            economic_code=economic_code,
            description=str(cell(row_idx, "description") or "").strip(),
            direction=direction,
        )

        # Importes según dirección
        if direction == "expense":
            line.initial_credits       = _to_decimal(cell(row_idx, "initial_credits"))
            line.modifications         = _to_decimal(cell(row_idx, "modifications"))
            line.final_credits         = _to_decimal(cell(row_idx, "final_credits"))
            line.commitments           = _to_decimal(cell(row_idx, "commitments"))
            line.recognized_obligations = _to_decimal(cell(row_idx, "recognized_obligations"))
            line.payments_made         = _to_decimal(cell(row_idx, "payments_made"))
            line.pending_payment       = _to_decimal(cell(row_idx, "pending_payment"))
        else:
            line.initial_forecast  = _to_decimal(cell(row_idx, "initial_forecast"))
            line.final_forecast    = _to_decimal(cell(row_idx, "final_forecast"))
            line.recognized_rights = _to_decimal(cell(row_idx, "recognized_rights"))
            line.net_collection    = _to_decimal(cell(row_idx, "net_collection"))
            line.pending_collection = _to_decimal(cell(row_idx, "pending_collection"))

            # Algunos XLSX de ingresos tienen columna modificaciones también
            line.modifications = _to_decimal(cell(row_idx, "modifications"))

        result.lines.append(line)
        result.total_rows_read += 1

    wb.close()

    logger.info(
        "xlsx_parse_done",
        file=path.name,
        lines=result.total_rows_read,
        skipped=result.total_rows_skipped,
        warnings=len(result.warnings),
    )
    return result
