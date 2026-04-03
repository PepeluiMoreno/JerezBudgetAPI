"""
Tests unitarios del parser XLSX.
No necesitan base de datos — trabajan con ficheros sintéticos en memoria.
"""
from __future__ import annotations

import io
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from etl.parsers.xlsx_execution import (
    ParsedBudgetLine,
    parse_execution_xlsx,
    _to_decimal,
    _clean_code,
    _is_subtotal_row,
)


# ── Helpers para crear XLSX de prueba ────────────────────────────────────────

def _make_expense_xlsx(rows: list[list]) -> Path:
    """Crea un XLSX de gastos sintético y lo guarda en un fichero temporal."""
    import tempfile
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ejecucion"

    # Fila 1: título (debe ser saltada)
    ws.append(["ESTADO DE EJECUCIÓN DE GASTOS 2026"])
    ws.append([])  # fila vacía

    # Fila 3: cabecera
    ws.append([
        "Orgánica", "Programa", "Económica", "Denominación",
        "Créditos Iniciales", "Modificaciones", "Créditos Definitivos",
        "A/D", "Obligaciones Reconocidas", "Pagos Realizados", "Pendiente de Pago",
    ])

    # Filas de datos
    for row in rows:
        ws.append(row)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return Path(tmp.name)


def _make_revenue_xlsx(rows: list[list]) -> Path:
    import tempfile
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["ESTADO DE EJECUCIÓN DE INGRESOS 2026"])
    ws.append([])
    ws.append([
        "Orgánica", "Económica", "Denominación",
        "Previsiones Iniciales", "Modificaciones", "Previsiones Definitivas",
        "Derechos Reconocidos Netos", "Recaudación Neta", "Pendiente de Cobro",
    ])

    for row in rows:
        ws.append(row)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return Path(tmp.name)


# ── Tests ────────────────────────────────────────────────────────────────────

class TestToDecimal:
    def test_integer(self):
        assert _to_decimal(1000) == Decimal("1000")

    def test_float(self):
        assert _to_decimal(1234.56) == Decimal("1234.56")

    def test_spanish_format(self):
        assert _to_decimal("1.234.567,89") == Decimal("1234567.89")

    def test_simple_comma(self):
        assert _to_decimal("1234,56") == Decimal("1234.56")

    def test_none(self):
        assert _to_decimal(None) is None

    def test_empty_string(self):
        assert _to_decimal("") is None

    def test_non_numeric(self):
        assert _to_decimal("N/A") is None

    def test_zero(self):
        assert _to_decimal(0) == Decimal("0")

    def test_negative(self):
        assert _to_decimal("-500,00") == Decimal("-500.00")


class TestCleanCode:
    def test_numeric_string(self):
        assert _clean_code("12000") == "12000"

    def test_float_artifact(self):
        assert _clean_code("1.0") == "1"

    def test_none(self):
        assert _clean_code(None) == ""

    def test_strips_whitespace(self):
        assert _clean_code("  0101  ") == "0101"


class TestSubtotalDetection:
    def test_total_row(self):
        assert _is_subtotal_row(["Total Capítulo 1", None, None, 500000]) is True

    def test_normal_row(self):
        assert _is_subtotal_row(["0101", "912", "12000", "Sueldos", 100000]) is False

    def test_suma_keyword(self):
        assert _is_subtotal_row(["Suma artículo", None, None, 200000]) is True


class TestParseExpenseXlsx:
    def test_basic_parse(self, tmp_path):
        path = _make_expense_xlsx([
            ["0101", "912", "12000", "Sueldos y salarios", 1_000_000, 50_000, 1_050_000,
             900_000, 850_000, 800_000, 50_000],
            ["0101", "912", "16000", "Cuotas Seg. Social", 400_000, 0, 400_000,
             350_000, 340_000, 330_000, 10_000],
        ])
        result = parse_execution_xlsx(path)
        path.unlink(missing_ok=True)

        assert result.direction == "expense"
        assert len(result.lines) == 2
        assert result.total_rows_read == 2
        assert result.total_rows_skipped >= 2  # título + vacía

        line = result.lines[0]
        assert line.economic_code == "12000"
        assert line.functional_code == "912"
        assert line.organic_code == "0101"
        assert line.initial_credits == Decimal("1000000")
        assert line.modifications == Decimal("50000")
        assert line.final_credits == Decimal("1050000")
        assert line.recognized_obligations == Decimal("850000")

    def test_skips_subtotals(self, tmp_path):
        path = _make_expense_xlsx([
            ["0101", "912", "12000", "Sueldos", 500_000, 0, 500_000, 450_000, 430_000, 420_000, 10_000],
            [None, None, None, "Total Capítulo 1", 500_000, 0, 500_000, 450_000, 430_000, 420_000, 10_000],
        ])
        result = parse_execution_xlsx(path)
        path.unlink(missing_ok=True)

        assert len(result.lines) == 1

    def test_empty_xlsx_returns_empty(self, tmp_path):
        import tempfile
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Sin datos relevantes"])
        f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(f.name)
        f.close()

        result = parse_execution_xlsx(Path(f.name))
        Path(f.name).unlink(missing_ok=True)

        assert len(result.lines) == 0
        assert len(result.warnings) > 0

    def test_hint_direction_revenue(self, tmp_path):
        path = _make_revenue_xlsx([
            ["0101", "11000", "IBI Urbana", 80_000_000, 0, 80_000_000, 75_000_000, 70_000_000, 5_000_000],
        ])
        result = parse_execution_xlsx(path, hint_direction="revenue")
        path.unlink(missing_ok=True)

        assert result.direction == "revenue"
        assert len(result.lines) == 1
        line = result.lines[0]
        assert line.initial_forecast == Decimal("80000000")
        assert line.recognized_rights == Decimal("75000000")

    def test_execution_rate_property(self, tmp_path):
        path = _make_expense_xlsx([
            ["0101", "912", "12000", "Sueldos", 1_000_000, 0, 1_000_000,
             900_000, 900_000, 850_000, 50_000],
        ])
        result = parse_execution_xlsx(path)
        path.unlink(missing_ok=True)

        line = result.lines[0]
        # execution_rate es propiedad del ORM, no del ParsedBudgetLine
        # Verificamos que los valores son correctos para calcularlo:
        assert line.recognized_obligations == Decimal("900000")
        assert line.final_credits == Decimal("1000000")
